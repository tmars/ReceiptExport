#coding=utf8
import sys
import openpyxl
import xlrd

def find_columns(ws, labels, from_row = 1):
	for row in xrange(from_row,from_row+MT_ROW_LIMIT):
		col_inds = {}
		for col in xrange(1,MT_COL_LIMIT):
			cell = ws.cell(row = row, column = col)
			if cell.value in [None, ""] or cell.data_type != 's':
				continue
			value = unicode(cell.value).lower()
			for l in set(labels) - set(col_inds.keys()):
				for text in MT_LABELS[l]:
					if text in value:
						col_inds[l] = col
		if len(col_inds.keys()) == len(labels):
			return (row, col_inds)
	return False

# возвращает индекс первой строки удовлетворяющей условиям
def find_row(ws, from_row, col_inds, funcs, row_count=1):
	prev_row = from_row -1
	rcount = 0
	for row in xrange(from_row, from_row+MT_ROW_LIMIT2):
		ret = True
		for l in col_inds:
			if not funcs[l](ws.cell(row = row, column = col_inds[l])):
				ret = False
				continue
		
		if ret: 
			if rcount == row_count-1:
				return row
			elif prev_row == (row-1):
				rcount += 1
			else:
				rcount = 0
			prev_row = row
	
	return False

def open_xls_as_xlsx(filename):
	# first open using xlrd
	book = xlrd.open_workbook(filename)
	index = 0
	nrows, ncols = 0, 0
	while nrows * ncols == 0:
		sheet = book.sheet_by_index(index)
		nrows = sheet.nrows
		ncols = sheet.ncols
		index += 1

	# prepare a xlsx sheet
	book1 = openpyxl.Workbook()
	sheet1 = book1.get_active_sheet()
	
	for row in xrange(0, nrows):
		for col in xrange(0, ncols):
			sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

	return book1

# возвращает список товаров
def get_products(filename):
	# открываем файл
	if filename[-4:] == '.xls':
		wb = open_xls_as_xlsx(filename)
	else:
		wb = openpyxl.load_workbook(filename)
	ws = wb.active
	
	# получаем номер строки и столбцов с заголовками
	lrow_ind, col_inds = find_columns(ws, ["price", "count", "name"])
	
	# получаем номер первой и последней строки с товарам
	srow_ind = find_row(ws, lrow_ind+1, col_inds, MT_FUNCS_NOCLEAN)
	frow_ind = find_row(ws, srow_ind+1, col_inds, MT_FUNCS_CLEAN)
	
	# формируем данные
	products = []
	for r in xrange(srow_ind, frow_ind):
		products.append({l: ws.cell(row = r, column = col_inds[l]).value for l in col_inds})
	
	print 'products',len(products)
	return products

# сохранение списка товаров в файл
def save_products(filename, products):
	# открываем файл
	wb = openpyxl.load_workbook(filename,guess_types=False,keep_vba=True)
	ws = wb.active
	
	# получаем номер строки и столбцов с заголовками
	lrow_ind, col_inds = find_columns(ws, ["price", "count", "name", "sum"])
	print 'ab',lrow_ind, col_inds
	sum_col_ind = col_inds["sum"]
	del col_inds["sum"]

	# получаем номер последней строки после товаров
	row_ind = find_row(ws, lrow_ind+1, 	col_inds, MT_FUNCS_CLEAN, row_count=2)
	print 'sd',row_ind
	row_ind -= 1
	
	# записываем данные
	first_row_ind = row_ind
	for prod in products:
		for key in prod:
			print 'r', row_ind, 'c', col_inds[key]
			ws.cell(row = row_ind, column = col_inds[key]).value = prod[key]
		row_ind += 1

	# подбиваем сумму
	letter = openpyxl.cell.get_column_letter(sum_col_ind)
	cell = ws.cell(row = row_ind, column = sum_col_ind)
	cell.value = "=SUBTOTAL(109,%s%s:%s%s)" % (letter, first_row_ind, letter, row_ind-1) 
	
	# для защищенных полей обязательно почистить атрибуты
	ws.formula_attributes[cell.coordinate] = {}
	
	# сохраняем в файл
	wb.save(filename+'_new.xlsx')

def is_zero(cell):
	return cell.value in ["", None, 0, "0"]
def is_digit(cell):
	return cell.data_type == 'n'
def is_string(cell):
	return cell.data_type == 's' and cell.value != ""

# область поиска заголовков
MT_COL_LIMIT = 100
MT_ROW_LIMIT = 100

# область поиска строк
MT_ROW_LIMIT2 = 50000

MT_FUNCS_CLEAN = {
	"name": is_zero,
	"count": is_zero,
	"price": is_zero,
}

MT_FUNCS_NOCLEAN = {
	"name": is_string,
	"count": is_digit,
	"price": is_digit,
}

MT_LABELS = {
	"name": [u"товар", u"название", u"наименование"],
	"count": [u"количество", u"кол-во", u"колво", u"кол"],
	"price": [u"цена"],
	"sum": [u"сумма", u"сум"],
}
